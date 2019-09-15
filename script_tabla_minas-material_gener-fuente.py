import openpyxl

proyectos_wb = openpyxl.load_workbook("Excel/datos_proyectos_excel.xlsx")
p_ws = proyectos_wb.active

# Tipos: 1 = Mina; 2 = Electrica; 3 = Vertedero

l_minas = list()
l_gener = list()
i = 2
while i < p_ws.max_row:
    tipo = p_ws[f"A{i}"].value
    nombre = p_ws[f"B{i}"].value
    if tipo == 1:
        mineral = p_ws[f"I{i}"].value
        l_minas.append((nombre, mineral))
    elif tipo == 2:
        fuente = p_ws[f"J{i}"].value
        l_gener.append((nombre, fuente))
    i += 1


minas_wb = openpyxl.load_workbook("Tablas/minas_mineral.xlsx")
minas_ws = minas_wb.active

minas_ws.cell(1, 1, "nombre")
minas_ws.cell(1, 2, "mineral")

i = 2
while l_minas:
    nombre, mineral = l_minas.pop()
    minas_ws.cell(i, 1, nombre)
    minas_ws.cell(i, 2, mineral)
    i += 1

gener_wb = openpyxl.load_workbook("Tablas/gener_fuente.xlsx")
gener_ws = gener_wb.active

gener_ws.cell(1, 1, "nombre")
gener_ws.cell(1, 2, "fuente")

i = 2
while l_gener:
    nombre, fuente = l_gener.pop()
    gener_ws.cell(i, 1, nombre)
    gener_ws.cell(i, 2, fuente)
    i += 1

gener_wb.save("Tablas/gener_fuente.xlsx")
minas_wb.save("Tablas/minas_mineral.xlsx")